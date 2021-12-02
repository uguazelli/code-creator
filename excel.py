import json
from openpyxl import load_workbook
import pandas as pd
from io import BytesIO
from flask import Blueprint, send_file, request
from util import generate_in_memory, unique_name_path

excel = Blueprint('excel','__name__')

@excel.route("/excel-to-json", methods=["POST"])
def excel_to_json_conversion():
    result = {}
    wb = load_workbook(filename=request.files["file0"])

    for sheet in wb:
        m_row = sheet.max_row
        m_col = sheet.max_column
        head = []
        line = []
        for row_index in range(1, m_row + 1):
            details = {}
            for col_index in range(1, m_col + 1):
                cell = sheet.cell(row=row_index, column=col_index).value
                if row_index == 1:
                    head.append(cell or "")
                else:
                    details[head[col_index - 1]] = cell
            if row_index != 1:
                line.append(details)
        result[sheet.title] = line

    return result

@excel.route("/json-to-excel", methods=["POST"])
def json_to_excel_conversion():
    req = request.get_json()
    obj = json.loads(req)
    df = pd.json_normalize(obj)
    # create an output stream
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine="xlsxwriter")
    # taken from the original question
    df.to_excel(writer, startrow=0, merge_cells=False, sheet_name="Sheet1", index=False)
    workbook = writer.book
    worksheet = writer.sheets["Sheet1"]
    format = workbook.add_format()
    format.set_bg_color("#eeeeee")
    worksheet.set_column(0, 9, 28)
    # the writer has done its job
    writer.close()
    # go back to the beginning of the stream
    output.seek(0)
    return output
    send_file(output, attachment_filename="result.xlsx", as_attachment=True)

@excel.route("/excel-to-csv", methods=["POST"])
def excel_to_csv_conversion():
    f = request.files["file0"]
    path = unique_name_path(f.filename)
    new_path = path.replace(".xlsx", ".csv")
    f.save(path)
    read_file = pd.read_excel(path)
    read_file.to_csv(new_path, index=None, header=True)
    return send_file(new_path, as_attachment=True)

@excel.route("/csv-to-excel", methods=["POST"])
def csv_to_excel_conversion():
    f = request.files["file0"]
    path = unique_name_path(f.filename)
    new_path = path.replace(".csv", ".xlsx")
    f.save(path)
    read_file = pd.read_csv(path)
    read_file.to_excel(new_path, index=None, header=True)
    return send_file(new_path, as_attachment=True)
