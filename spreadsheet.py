from openpyxl import load_workbook
import pandas as pd
from io import BytesIO
from flask import current_app, send_file
from util import generate_in_memory, unique_name
import os


def excel_to_json_conversion(filename):
    result = {}
    wb = load_workbook(filename=filename)

    for sheet in wb:
        m_row = sheet.max_row
        m_col = sheet.max_column
        head = []
        line = []
        for row_index in range(1, m_row + 1):
            details = {}
            for col_index in range(1, m_col + 1):
                cell = sheet.cell(row=row_index, column=col_index).value
                if row_index == 1:
                    head.append(cell or "")
                else:
                    details[head[col_index - 1]] = cell
            if row_index != 1:
                line.append(details)
        result[sheet.title] = line

    return result


def json_to_spreadsheet(json):
    df = pd.json_normalize(json)

    # create an output stream
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine="xlsxwriter")

    # taken from the original question
    df.to_excel(writer, startrow=0, merge_cells=False, sheet_name="Sheet1", index=False)
    workbook = writer.book
    worksheet = writer.sheets["Sheet1"]
    format = workbook.add_format()
    format.set_bg_color("#eeeeee")
    worksheet.set_column(0, 9, 28)

    # the writer has done its job
    writer.close()

    # go back to the beginning of the stream
    output.seek(0)

    return output


# def excel_to_csv_conversion(request):
#    f = request.files["file"]
#    path = unique_name(f.filename)
#    f.save(path)
#    read_file = pd.read_excel(path)
#    read_file.to_csv(path, index=None, header=True)
#
#    r = current_app.response_class(generate_in_memory(path), mimetype="text/csv")
#    r.headers.set("Content-Disposition", "attachment", filename="result.csv")
#    return r


def excel_to_csv_conversion(request):
    f = request.files["file"]
    path = unique_name(f.filename)
    new_path = path.replace(".xlsx", ".csv")
    f.save(path)
    read_file = pd.read_excel(path)
    read_file.to_csv(new_path, index=None, header=True)

    return send_file(
        new_path,
        as_attachment=True,
    )


def csv_to_excel_conversion(request):
    f = request.files["file"]
    path = unique_name(f.filename)
    new_path = path.replace(".csv", ".xlsx")
    f.save(path)
    read_file = pd.read_csv(path)
    read_file.to_excel(new_path, index=None, header=True)

    return send_file(
        new_path,
        as_attachment=True,
    )
