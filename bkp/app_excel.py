import json
from openpyxl import load_workbook
import pandas as pd
from io import BytesIO
from flask import Blueprint, send_file, request, render_template
from util import generate_in_memory, unique_name_path

app_excel = Blueprint('app_excel','__name__')

@app_excel.route("/excel-to-json", methods=["GET","POST"])
def excel_to_json_conversion():
    if request.method == "GET":
        return render_template('convert.html', title="Excel to Json", js_param='', js_file='js/excel_to_json.js', multiple='')
    else:
        result = {}
        wb = load_workbook(filename=request.files["file0"])

        for sheet in wb:
            m_row = sheet.max_row
            m_col = sheet.max_column
            head = []
            line = []
            for row_index in range(1, m_row + 1):
                details = {}
                for col_index in range(1, m_col + 1):
                    cell = sheet.cell(row=row_index, column=col_index).value
                    if row_index == 1:
                        head.append(cell or "")
                    else:
                        details[head[col_index - 1]] = cell
                if row_index != 1:
                    line.append(details)
            result[sheet.title] = line

        return result


@app_excel.route("/json-to-excel", methods=["GET", "POST"])
def json_to_excel_conversion():
    if request.method == "GET":
        return render_template('convert.html', title="Json to Excel", js_param='', js_file='js/json_to_excel.js', multiple='')
    else:
        req = request.get_json()
        obj = json.loads(req)
        df = pd.json_normalize(obj)
        # create an output stream
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine="xlsxwriter")
        # taken from the original question
        df.to_excel(writer, startrow=0, merge_cells=False, sheet_name="Sheet1", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]
        format = workbook.add_format()
        format.set_bg_color("#eeeeee")
        worksheet.set_column(0, 9, 28)
        # the writer has done its job
        writer.close()
        # go back to the beginning of the stream
        output.seek(0)
        send_file(output, attachment_filename="result.xlsx", as_attachment=True)


@app_excel.route("/excel-to-csv", methods=["GET","POST"])
def excel_to_csv_conversion():
    if request.method == "GET":
        return render_template('convert.html', title="Excel to CSV", js_param='excelToCSV', js_file='js/Controller.js', multiple='')
    else:
        f = request.files["file0"]
        path = unique_name_path(f.filename)
        new_path = path.replace(".xlsx", ".csv")
        f.save(path)
        read_file = pd.read_excel(path)
        read_file.to_csv(new_path, index=None, header=True)
        return send_file(new_path, as_attachment=True)


@app_excel.route("/csv-to-excel", methods=["GET","POST"])
def csv_to_excel_conversion():
    if request.method == "GET":
        return render_template('convert.html', title="CSV to Excel", js_param='csvToExcel', js_file='js/Controller.js', multiple='')
    else:
        f = request.files["file0"]
        path = unique_name_path(f.filename)
        new_path = path.replace(".csv", ".xlsx")
        f.save(path)
        read_file = pd.read_csv(path)
        read_file.to_excel(new_path, index=None, header=True)
        return send_file(new_path, as_attachment=True)
